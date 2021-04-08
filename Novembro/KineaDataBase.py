import pandas as pd
from openpyxl import load_workbook
from time import process_time 
from datetime import timedelta
import pyodbc
import calendar

t1_start = process_time() 

#CONECTANDO COM O SQL SERVER
server = 'USER-PC' 
database = 'KineaBandeira' 
username = '' 
password = '' 

cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';Trusted_Connection=yes')
cursor = cnxn.cursor()

#Arquivos para Ler
wb = load_workbook(filename = "planilhaPrincipal.xlsx", data_only=True)
wb_Fechamento = load_workbook(filename = "planilhaIncorporadoraConverted.xlsx", data_only=True)

#Sheets
sheet_FluxoInvestidores = wb['Fluxo Investidores']
sheet_Fluxo = wb['Fluxo Projeto']
sheet_Premissas = wb['Premissas']
sheet_Unidades = wb['Tabela Vendas']
sheet_Classificacao = wb['Tabela Vendas']
sheet_Tipo = wb['Tabela Vendas']
sheet_Blocos = wb['Tabela Vendas']
sheet_Andar = wb['Tabela Vendas']
sheet_Inadimplencia = wb_Fechamento['INADIMPLÊNCIA']
sheet_LTV = wb['M-POS. FIN.']
sheet_Obra = wb['Obra']
sheet_BandeiraCo = wb['Fluxo Bandeira Co']
sheet_Curvas = wb['Curvas']
sheet_DRE = wb['DRE']

###Listas para acrescentar nos DataFrames necessários###

#Listas dos Investidores
lista_Investidores = []
lista_titulos_investidores = ['DataInvestimento','SCP1','SCP2','SCP3','SCP4','SCP5','SCP6','SCP7','SCP8',
                              'SCP9','SCP10','SCP11','SCP12','SCP13']
lista_Investidores_colunas = [1,12,13,14,15,16,17,18,19,20,21,22,23,24]
lista_Investidores_Titulos_Index = ['ID_Projeto','ID_Fluxo']

#Listas do Fluxo
lista_Fluxo = []
lista_Titulos_Fluxo = ['Data','Vendido','AVender','PermutaKinea','PermutaLDI','Impostos','CustodeObra',
                       'TaxaAdminObra','Carrego','Incorporacao','ComercialMarketing',
                       'TaxaGestaoImob','OutrasDespesas','OutrasReceitas','Terreno',
                       'CEPAC','RecDespFinanceiras','Aportes','Distribuicoes','CaixaFinalEfetivo',
                       'SaldoNominal','SaldoAcumulado','SaldoReal']
lista_Fluxo_Colunas = [1,7,8,9,10,13,14,15,16,17,18,19,20,21,27,28,32,36,37,41,43,44,46]
lista_Fluxo_Titulos_Index = ['BaseData', 'ID_Projeto', 'ID_Fluxo']
lista_Fluxo_Colunas_Curvas = [5,6]
lista_Fluxo_Titulos_Curvas = ['PercentualMesVendas', 'PercentualAcumuladoVendas']

#Listas dos Indicadores Gerais
lista_IndicadoresGerais_Titulos_Index = ['DataBaseTerreno', 'DataBaseObra', 'ValorMetroTerreno', 'ValorMetroObra','AreadoTerreno', 'ValordoTerreno','ID_Fluxo']
lista_ValorMetroObra_Colunas = [4]
list_IndicadoresGerais_Titulos_Index2 = ['ID_Projeto']

#Listas das Unidades
lista_Unidades = []
lista_Titulos_Unidades = ['ID_Andar','Área','Status','NúmeroUnidade','DataVenda','MesVenda','ValorVendidoOriginal','PreçoPorMetroQVendido']
lista_Unidades_Colunas = [3,4,5,2,6,7,8,9]
lista_Unidades_Titulos_Index = ['ID_Unidade', 'ID_Tipo', 'ID_Classificacao']
lista_Unidades_Colunas_Index = [2]

#Listas de Classificação
lista_Classificacao = []
lista_Classificacao_Colunas = [4]

#Listas dos Blocos
lista_Blocos = []
lista_Titulos_Blocos = ['NomeBloco']
lista_Blocos_Colunas = [1]
lista_Blocos_Titulos_Index = ['ID_Projeto', 'ID_Bloco']

#Listas dos Andar
lista_Andar = []
lista_Titulos_Andar = ['Andar']
lista_Andar_Colunas = [3]
lista_Andar_Titulos_Index = ['ID_Bloco', 'ID_Andar']

#Listas dos Indicadores Mensais
lista_IndicadoresMensais_Titulos_Index1 = ['ID_Projeto']
lista_IndicadoresMensais_Titulos_Index3 = ['ID_Fluxo']
lista_IndicadoresMensais = []
lista_Titulos_IndicadoresMensais = ['Andar']
lista_IndicadoresMensais_Colunas = [3]
lista_IndicadoresMensais_Titulos_Index2 = ['TIRReal', 'TIRNominal','Inadimplência','ValorMetroHistórico',
                                          'ValorMetroEstoque', 'Lucro', 'LTV', 'PercentualAvançoObra', 
                                          'BD_Incorporacao_Fluxo_Data','Caixa','Receita','ReceitaPermuta',
                                          'Impostos','CustodeObra','TaxaAdminObra','Carrego','Incorporacao',
                                          'Marketing','TaxaGestaoImob', 'OutrasDespesas','TotalDespesas',
                                          'TerrenoCEPAC','RecDespFinanceiras','NAVTotal','NAVSCP','TIRRealProjeto',
                                          'TIRNominalProjeto','TIRRealHolding','TIRRealHoldingPosPerf',
                                          'TIRNominalHolding','Multiplo','Performance']
lista_Inadimplencia = []
lista_Inadimplencia_Colunas = [8, 9, 12]
lista_ValorMetroHist_Colunas = [9]
lista_ValorMetroEstoque_Colunas = [17]
lista_LTV_Colunas = [15]
lista_PercentualAvancoObra_Colunas = [30]

#Listas dos Investidores Bandeira CO

lista_BandeiraCo = []
lista_Titulos_BandeiraCo = ['Data','ChamadaEfetiva','DistribuicaoEfetiva','ChamadaEquity','ChamadaPermuta','SPEBandeiraIncorporacaoEquity',
                       'SPEBandeiraIncorporacaoPermuta','SPEBandeiraIncorporacao','Impostos','Auditoria','MonitoramentoObra',
                       'Juridico','TaxaAdministracao','TaxaPerfomancePermutas','Contabilidade','OutrasDespesas',
                       'OutrasReceitas','SaldoNominal','SaldoAcumulado','SaldoReal','SaldoRealAcumulado',
                       'BasePerformance','Performance','SaldoNominalPosPerformance','SaldoAcumuladoNominalPosPerformance','SaldoRealPosPerformance']
lista_BandeiraCo_Colunas = [1,6,7,9,10,12,13,14,19,20,21,22,23,24,25,26,27,44,45,47,49,50,51,52,53,55]
lista_BandeiraCo_Titulos_Index = ['ID_Projeto','ID_Fluxo','BaseData']

###Fim das Listas###


###FUNÇÕES PARA LER E TRATAR OS DADOS###
def UniqueList(sheet, tipo, row, col, name_col, name_table):

    if sheet == sheet_FluxoInvestidores or sheet == sheet_Fluxo or sheet == sheet_BandeiraCo:
        date_start = sheet.cell(row = 2, column = 1).value
        date_end = sheet.cell(row = 86, column = 1).value
        meses = date_end.month - date_start.month
        anos = (date_end.year - date_start.year) * 12
        n = meses+anos+1
    elif sheet == sheet_Premissas:
        n = 1
    elif sheet == sheet_Blocos or sheet == sheet_Andar:
        n = 46
    
    listofzeros = [0] * n
     
    #PARA LER O INDEX
    getindex = pd.read_sql_query(
    '''SELECT
    {}
    FROM {}'''.format(name_col, name_table), cnxn)
    df = pd.DataFrame(getindex, columns=['{}'.format(name_col)])
    
    if tipo == "index" and row == 0 and col == 0: 
        if df.empty == False:
            index_base_atual = df.iloc[-1]["{}".format(name_col)]
        else:
            index_base_atual = 0
        index_base_atual += 1
        for i in range (len(listofzeros)):
            listofzeros[i] = index_base_atual
     #PARA LER A BASEDATA
    elif tipo == "data" and row == 0 and col == 0: 
        if df.empty == False:
            index_base_atual = df.iloc[-1]["{}".format(name_col)]
            days_in_month = calendar.monthrange(index_base_atual.year, index_base_atual.month)[1]
            for i in range (len(listofzeros)):     
                listofzeros[i] = index_base_atual + timedelta(days = days_in_month)
        else:
            index_base_atual = sheet_Fluxo.cell(row = 2, column = 1).value
            for i in range (len(listofzeros)):     
                listofzeros[i] = index_base_atual
     #PARA LER A CELL    
    elif tipo == "cell" and row != 0 and col != 0:  
        uniqueCell = sheet.cell(row = row, column = col).value
        for i in range(len(listofzeros)):
            listofzeros[i] = uniqueCell 
            
    else:
        print("Tipo inválido: escolha entre index, data ou cell")
  
    list_index = listofzeros
    
    return list_index 

def readExcelColumn(sheet, lista, row, columns):
    
    if sheet == sheet_Classificacao or sheet == sheet_Andar:
        m_row = 47
    elif sheet == sheet_Inadimplencia:
        m_row = sheet_Inadimplencia.max_row
    elif sheet == sheet_LTV:
        m_row = 27
    elif sheet == sheet_Obra:
        m_row = 85
    else:
        m_row = 86
          
    for cols in columns:
        for col in sheet.iter_cols(min_row = row, min_col = cols, max_row = m_row, max_col = cols):
            for cell in col:
                lista.append(cell.value)

    return lista

def AddDataFrame(lista, lista_de_titulos):
    
    lst1 = [0] * (len(lista)//len(lista_de_titulos))
    df = pd.DataFrame(
    {'{}'. format("ColunaAgregação"): lst1})
    
    lista_separada = [0] * (len(lista)//len(lista_de_titulos))
    cont = 0
    
    for index in range(len(lista_de_titulos)):
        for i in range (len(lista)//len(lista_de_titulos)):  
            lista_separada[i] = lista[cont]
            cont += 1
            df['{}'.format(lista_de_titulos[index])] = lista_separada

    return df

def AddDataFrameColumn(df, lista, lista_de_titulos_index, index):

    df['{}'.format(lista_de_titulos_index[index])] = lista

    return


#CRIANDO DATA FRAMES PARA CADA UMA DAS TABELAS
def DataFrame_Investidores():

    df_Investidores = AddDataFrame(readExcelColumn(sheet_FluxoInvestidores, lista_Investidores, 2, lista_Investidores_colunas), lista_titulos_investidores)
    
    lista1 = UniqueList(sheet_FluxoInvestidores, "index", 0, 0, "ID_Projeto", "BD_Incorporacao_Investidores")
    for i in range(len(lista1)):  #POIS ESTAMOS USANDO SOMENTE UM PROJETO
        lista1[i] = 1
    AddDataFrameColumn(df_Investidores, lista1 , lista_Investidores_Titulos_Index, 0)
    AddDataFrameColumn(df_Investidores, UniqueList(sheet_FluxoInvestidores, "index", 0, 0, "ID_Fluxo", "BD_Incorporacao_Investidores"), lista_Investidores_Titulos_Index, 1)

    return df_Investidores

def DataFrame_Fluxo():
    
    df_Fluxo = AddDataFrame(readExcelColumn(sheet_Fluxo, lista_Fluxo, 2, lista_Fluxo_Colunas), lista_Titulos_Fluxo)
    lista_Curvas1 = [0]*85
    cont = 4
    for i in range (36, 85):
        lista_Curvas1[i] = round(sheet_Curvas.cell(row=cont,column=5).value,5)
        cont += 1
    lista_Curvas2 = [0]*85
    cont = 4
    for i in range (36, 85):
        lista_Curvas2[i] = round(sheet_Curvas.cell(row=cont,column=6).value,5)
        cont += 1
    
    lista1 = UniqueList(sheet_Fluxo, "index", 0, 0, "ID_Projeto", "BD_Incorporacao_Fluxo")
    for i in range(len(lista1)):  #POIS ESTAMOS USANDO SOMENTE UM PROJETO
        lista1[i] = 1
    
    AddDataFrameColumn(df_Fluxo, lista_Curvas1, lista_Fluxo_Titulos_Curvas, 0)
    AddDataFrameColumn(df_Fluxo, lista_Curvas2, lista_Fluxo_Titulos_Curvas, 1)        
    AddDataFrameColumn(df_Fluxo, lista1, lista_Fluxo_Titulos_Index, 1)
    
    #A função AddDataFrameColumn abaixo só funciona se a basedata da planilha for a posterior a basedata da base de dados
    #AddDataFrameColumn(df_Fluxo, UniqueList(sheet_Fluxo, "data", 0, 0, "BaseData", "BD_Incorporacao_Fluxo"), lista_Fluxo_Titulos_Index, 0)
    #Caso a função acima não se encaixe, a função abaixo se encaixará, mas a planilha deverá estar correta
    lista_data = []
    for i in range (85):
        lista_data.append(sheet_Premissas.cell(row=2,column=3).value)
    AddDataFrameColumn(df_Fluxo, lista_data, lista_Fluxo_Titulos_Index, 0)
    
    AddDataFrameColumn(df_Fluxo, UniqueList(sheet_Fluxo, "index", 0, 0, "ID_Fluxo", "BD_Incorporacao_Fluxo"), lista_Fluxo_Titulos_Index, 2)
    
    df_Fluxo = df_Fluxo.fillna(0)
    
    return df_Fluxo

def DataFrame_IndicadoresGerais():
    #ID_Projeto
    df_IndicadoresGerais = AddDataFrame(UniqueList(sheet_Premissas, "index", 0, 0, "ID_Projeto", "BD_Incorporacao_IndicadoresGerais"), lista_IndicadoresMensais_Titulos_Index1) 
    #Área do Terreno
    lista_AreaTerreno = [1333]
    AddDataFrameColumn(df_IndicadoresGerais, lista_AreaTerreno, lista_IndicadoresGerais_Titulos_Index, 4)
    #Valor do Terreno
    lista_ValorTerreno = [21326358.96]
    AddDataFrameColumn(df_IndicadoresGerais, lista_ValorTerreno, lista_IndicadoresGerais_Titulos_Index, 5)
    #DataBaseTerreno
    lista_DataBaseTerreno = [sheet_Premissas.cell(row=7, column = 3).value]
    AddDataFrameColumn(df_IndicadoresGerais, lista_DataBaseTerreno, lista_IndicadoresGerais_Titulos_Index, 0)
    #DataBaseObra
    lista_DataBaseTerreno = [sheet_Premissas.cell(row=9, column = 3).value]
    AddDataFrameColumn(df_IndicadoresGerais, lista_DataBaseTerreno, lista_IndicadoresGerais_Titulos_Index, 1)
    #ValorMetroTerreno
    lista_ValorMetroTerreno = [lista_ValorTerreno[0]/lista_AreaTerreno[0]]
    AddDataFrameColumn(df_IndicadoresGerais, lista_ValorMetroTerreno, lista_IndicadoresGerais_Titulos_Index, 2)
    #ValorMetroObra
    lista1 = []
    lista_calc_INCC = UniqueList(sheet_Premissas, "index", 0, 0, "ID_Fluxo", "BD_Incorporacao_IndicadoresGerais")
    AddDataFrameColumn(df_IndicadoresGerais, lista_calc_INCC, lista_IndicadoresGerais_Titulos_Index, 6)
    readExcelColumn(sheet_Classificacao, lista1, 2, lista_ValorMetroObra_Colunas)
    ValorMetroObra = (sheet_Obra.cell(row=13, column = 3).value)/(sum(lista1))
    INCC = (sheet_Obra.cell(row = lista_calc_INCC[0]+1, column = 7).value)/(sheet_Obra.cell(row=9, column = 3).value)
    lista_ValorMetroObraINCC = [ValorMetroObra*INCC]
    AddDataFrameColumn(df_IndicadoresGerais, lista_ValorMetroObraINCC, lista_IndicadoresGerais_Titulos_Index, 3)   

    return df_IndicadoresGerais

def DataFrame_Unidades():
    #df_Unidades
    df_Unidades = AddDataFrame(readExcelColumn(sheet_Unidades, lista_Unidades, 2, lista_Unidades_Colunas), lista_Titulos_Unidades) 
    df_Unidades['ValorVendidoOriginal'] = df_Unidades['ValorVendidoOriginal'].replace(['-'],0)
    df_Unidades['PreçoPorMetroQVendido'] = df_Unidades['PreçoPorMetroQVendido'].replace(['-'],0)
    df_Unidades['DataVenda'] = df_Unidades['DataVenda'].replace(['-'],'')
    df_Unidades['MesVenda'] = df_Unidades['MesVenda'].replace(['-'],'')    
    
    #Unidade
    cont = 0
    lista_ID_Unidade = []
    readExcelColumn(sheet_Classificacao, lista_ID_Unidade, 2, lista_Unidades_Colunas_Index)
    for i in range (len(lista_ID_Unidade)):
        cont += 1
        lista_ID_Unidade[i] = cont
    AddDataFrameColumn(df_Unidades, lista_ID_Unidade, lista_Unidades_Titulos_Index, 0)
    #Tipo
    list_id_tipo = [1]*len(lista_ID_Unidade) #1, pois o ID de Apartamentos é 1 e são todos Apartamentos 
    AddDataFrameColumn(df_Unidades, list_id_tipo, lista_Unidades_Titulos_Index, 1)
    #Classificacao
    lista = readExcelColumn(sheet_Classificacao, lista_Classificacao, 2, lista_Classificacao_Colunas)
    list_id_class = [0]*len(lista)
    for i in range (len(lista)):
        if lista[i] < 30:
            list_id_class[i] = 1
        elif lista[i] > 31 and lista[i] < 45:
            list_id_class[i] = 2
        elif lista[i] > 46 and lista[i] < 65:
            list_id_class[i] = 3            
        elif lista[i] > 66 and lista[i] < 85:
            list_id_class[i] = 4            
        elif lista[i] > 86 and lista[i] < 130:
            list_id_class[i] = 5            
        elif lista[i] > 131 and lista[i] < 180:
            list_id_class[i] = 6       
        elif lista[i] > 180:
            list_id_class[i] = 7   
    AddDataFrameColumn(df_Unidades, list_id_class, lista_Unidades_Titulos_Index, 2)

    return df_Unidades

def DataFrame_Bloco():
    
    lista1 = readExcelColumn(sheet_Blocos, lista_Blocos, 2, lista_Blocos_Colunas)
    lista2 = []
    lista3 = []
    listaID = UniqueList(sheet_Premissas, "index", 0, 0, "ID_Bloco", "BD_Incorporacao_Bloco")
    listaBloco = []
    for i in range(len(lista1)):
        lista2.append(lista1[i].split("-"))
        lista3.append(lista2[i][0]) 
    listaBloco.append(lista3[0])
    for i in range(len(lista3)):
        if lista3[i] != lista3[i-1]:
            listaID.append(listaID[-1]+1) 
            listaBloco.append(lista3[i])
        else:
            pass
    df_Bloco = AddDataFrame(listaBloco, lista_Titulos_Blocos)
    lista6 = UniqueList(sheet_Premissas, "index", 0, 0, "ID_Projeto", "BD_Incorporacao_Bloco")
    for i in range(len(lista6)):  #POIS ESTAMOS USANDO SOMENTE UM PROJETO
        lista6[i] = 1    
    AddDataFrameColumn(df_Bloco, lista6, lista_Blocos_Titulos_Index, 0)
    AddDataFrameColumn(df_Bloco, listaID, lista_Blocos_Titulos_Index, 1)
    
    return df_Bloco

def DataFrame_Andar():
    
    lista1 = readExcelColumn(sheet_Andar, lista_Andar, 2, lista_Andar_Colunas)    
    df_Andar = AddDataFrame(lista1, lista_Titulos_Andar)
    AddDataFrameColumn(df_Andar, UniqueList(sheet_Andar, "index", 0, 0, "ID_Bloco", "BD_Incorporacao_Andar"), lista_Andar_Titulos_Index, 0)
    for i in range(len(lista1)):
        lista1[i] = int(lista1[i])
    lista2 =[1]
    for i in range(1,len(lista1)):
        if lista1[i] == lista1[i-1]:
            lista2.append(lista1[i-1])
        else:
            lista2.append(lista2[-1]+1)
    AddDataFrameColumn(df_Andar, lista2, lista_Andar_Titulos_Index, 1)
    return df_Andar

def DataFrame_IndicadoresMensais():
    #ID_Projeto
    df_IndicadoresMensais = AddDataFrame(UniqueList(sheet_Premissas, "index", 0, 0, "ID_Fluxo", "BD_Incorporacao_IndicadoresMensais"), lista_IndicadoresMensais_Titulos_Index3)    
    #TIRReal
    lista_TIRReal = [sheet_Premissas.cell(row=18, column = 17).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_TIRReal, lista_IndicadoresMensais_Titulos_Index2, 0)
    #TIRNominal
    lista_TIRNominal = [sheet_Premissas.cell(row=21, column = 17).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_TIRNominal, lista_IndicadoresMensais_Titulos_Index2, 1)
    #Inadimplencia
    lista1 = readExcelColumn(sheet_Inadimplencia, lista_Inadimplencia, 7, lista_Inadimplencia_Colunas)
    lista_Inad = [sum(lista1)]
    AddDataFrameColumn(df_IndicadoresMensais, lista_Inad, lista_IndicadoresMensais_Titulos_Index2, 2)
    #ValorMetroHistorico
    lista2 = []
    cont = 0
    readExcelColumn(sheet_Classificacao, lista2, 2, lista_ValorMetroHist_Colunas)
    for i in range (len(lista2)):
        if lista2[i] == "-":
            lista2[i] = 0
        else:
            cont += 1
    listaMetroHist = [sum(lista2)/cont]
    AddDataFrameColumn(df_IndicadoresMensais, listaMetroHist, lista_IndicadoresMensais_Titulos_Index2, 3)
    #ValorMetroEstoque
    lista3 = []
    cont = 0
    readExcelColumn(sheet_Classificacao, lista3, 2, lista_ValorMetroEstoque_Colunas)
    for i in range (len(lista3)):
        if lista3[i] == "-":
            lista3[i] = 0
        else:
            cont += 1
    listaMetroEstoque = [(sum(lista3)/cont)/(1-sheet_Classificacao.cell(row=2, column = 13).value)]
    AddDataFrameColumn(df_IndicadoresMensais, listaMetroEstoque, lista_IndicadoresMensais_Titulos_Index2, 4)
    #Lucro
    lista_Lucro = [sheet_Premissas.cell(row=20, column = 17).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_Lucro, lista_IndicadoresMensais_Titulos_Index2, 5)    
    #LTV
    lista4 = []
    readExcelColumn(sheet_LTV, lista4, 7, lista_LTV_Colunas)
    lista_LTV = [sum(lista4)/(len(lista4)+sheet_Premissas.cell(row=21,column=11).value)]
    AddDataFrameColumn(df_IndicadoresMensais, lista_LTV, lista_IndicadoresMensais_Titulos_Index2, 6)
    #PercentualAvançoObra
    lista5 = []
    readExcelColumn(sheet_Obra, lista5, 2, lista_PercentualAvancoObra_Colunas)
    lista_Obra = [sum(lista5)]
    AddDataFrameColumn(df_IndicadoresMensais, lista_Obra, lista_IndicadoresMensais_Titulos_Index2, 7)
    
    #BD_Incorporacao_Fluxo_Data
    #Mesmo caso do database do Fluxo
    #AddDataFrameColumn(df_IndicadoresMensais, UniqueList(sheet_Premissas, "data", 0, 0, "BD_Incorporacao_Fluxo_Data", "BD_Incorporacao_IndicadoresMensais"), lista_IndicadoresMensais_Titulos_Index2, 8)
    #Caso o banco de dados não esteja em sintonia com a planilha usar a função abaixo
    lista_data = []
    lista_data.append(sheet_Premissas.cell(row=2,column=3).value)
    AddDataFrameColumn(df_IndicadoresMensais, lista_data, lista_IndicadoresMensais_Titulos_Index2, 8)
    #Performance
    lista_Performance = [sheet_Premissas.cell(row=38,column=14).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_Performance , lista_IndicadoresMensais_Titulos_Index2, 31)
    #DRE
    #Caixa
    lista_DRE = [sheet_DRE.cell(row=4,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 9)
    #Receita (Carteira + Estoque) - ex permuta
    lista_DRE = [sheet_DRE.cell(row=5,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 10)
    #Receita Permuta 
    lista_DRE = [sheet_DRE.cell(row=6,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 11)
    #Impostos
    lista_DRE = [sheet_DRE.cell(row=8,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 12)
    #Custo de Obra
    lista_DRE = [sheet_DRE.cell(row=9,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 13)
    #Taxa de Administração Obra
    lista_DRE = [sheet_DRE.cell(row=10,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 14)
    #Carrego
    lista_DRE = [sheet_DRE.cell(row=11,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 15)
    #Incorporação
    lista_DRE = [sheet_DRE.cell(row=12,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 16)
    #Marketing
    lista_DRE = [sheet_DRE.cell(row=13,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 17)   
    #Taxa de Gestão Imobiliária    
    lista_DRE = [sheet_DRE.cell(row=14,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 18)
    #Outras Despesas
    lista_DRE = [sheet_DRE.cell(row=15,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 19)
    #TOTAL DESPESAS
    lista_DRE = [sheet_DRE.cell(row=16,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 20)
    #Terreno + CEPAC
    lista_DRE = [sheet_DRE.cell(row=18,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 21)
    #Receitas/Despesas Financeiras
    lista_DRE = [sheet_DRE.cell(row=19,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 22)
    #NAV TOTAL (c/ permuta)
    lista_DRE = [sheet_DRE.cell(row=21,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 23)
    #NAV SCP
    lista_DRE = [sheet_DRE.cell(row=23,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 24)
    #TIR REAL PROJETO
    lista_DRE = [sheet_DRE.cell(row=25,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 25)
    #TIR NOMINAL PROJETO
    lista_DRE = [sheet_DRE.cell(row=26,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 26)      
    #TIR REAL HOLDING
    lista_DRE = [sheet_DRE.cell(row=28,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 27)
    #TIR REAL HOLDING - PÓS-PERF.
    lista_DRE = [sheet_DRE.cell(row=29,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 28)
    #TIR NOMINAL HOLDING
    lista_DRE = [sheet_DRE.cell(row=30,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 29)
    #MULTIPLO
    lista_DRE = [sheet_DRE.cell(row=31,column=4).value]
    AddDataFrameColumn(df_IndicadoresMensais, lista_DRE , lista_IndicadoresMensais_Titulos_Index2, 30)       
    
    return df_IndicadoresMensais       

def DataFrame_BandeiraCo():
    
    df_BandeiraCo = AddDataFrame(readExcelColumn(sheet_BandeiraCo, lista_BandeiraCo, 2, lista_BandeiraCo_Colunas), lista_Titulos_BandeiraCo)
    lista1 = UniqueList(sheet_BandeiraCo, "index", 0, 0, "ID_Projeto", "BD_Incorporacao_Investidores_BandeiraCo")
    for i in range(len(lista1)):  #POIS ESTAMOS USANDO SOMENTE UM PROJETO
        lista1[i] = 1
    AddDataFrameColumn(df_BandeiraCo, lista1, lista_BandeiraCo_Titulos_Index, 0)
    lista2 = UniqueList(sheet_BandeiraCo, "data", 0, 0, "BaseData", "BD_Incorporacao_Investidores_BandeiraCo")
    AddDataFrameColumn(df_BandeiraCo, lista2, lista_BandeiraCo_Titulos_Index, 2)
    AddDataFrameColumn(df_BandeiraCo, UniqueList(sheet_BandeiraCo, "index", 0, 0, "ID_Fluxo", "BD_Incorporacao_Investidores_BandeiraCo"), lista_BandeiraCo_Titulos_Index, 1)
    df_BandeiraCo = df_BandeiraCo.fillna(0)

    return df_BandeiraCo


###CRIAR TODOS OS DATAFRAMES###
def Update_DataBase():
    try:
        try:
            DataFrame_Investidores()
        except:
            print("Ocorreu um problema no dataframe Investidores") 
        try:
            DataFrame_Fluxo()
        except:
            print("Ocorreu um problema no dataframe Fluxo") 
        try:
            DataFrame_IndicadoresGerais()
        except:
            print("Ocorreu um problema no dataframe IndicadoresGerais") 
        try:
            DataFrame_Unidades()
        except:
            print("Ocorreu um problema no dataframe Unidades") 
        try:
            DataFrame_Bloco()
        except:
            print("Ocorreu um problema no dataframe Bloco") 
        try:
            DataFrame_Andar()
        except:
            print("Ocorreu um problema no dataframe Andar") 
        try:
            DataFrame_IndicadoresMensais()
        except:
            print("Ocorreu um problema no dataframe IndicadoresMensais") 
        try:
            DataFrame_BandeiraCo()
        except:
            print("Ocorreu um problema no dataframe BandeiraCO") 
        x = "Todos os dataframes foram criados corretamente!"
    except:
        x = "Ocorreu um problema na criação de dataframes!"
    return print('{}'.format(x))


###ENVIAR DATAFRAMES PARA O SERVIDOR###
def _TextoPMatriz(Texto):
            
    #1: Descobre quantos parâmetros
    Cont = 1
    for i in range(0,len(Texto),1):
        if( Texto[i:i+1] == ";"):
            Cont += 1

    if( Cont % 2 == 1 ):
        return False #Número tem que ser par!
    
    #Cria a Matriz
    Matriz = []
    Col = 1
    temp1, temp2 = '', ''

    for i in range(0,len(Texto)+1,1):

        if( Texto[i:i+1] == ";" or Texto[i:i+1] == ""): 
            Col += 1
            if( Col > 2 ):
                Col = 1
                Matriz.append((temp1,temp2))
                temp1, temp2 = '', ''
        else:
            if(Col == 1):
                temp1 += Texto[i:i+1] 
            else:
                temp2 += Texto[i:i+1] 

    return Matriz 

def Com_Soft_Add(cursor, Tabela, CamposValores):
        """
         Mesma coisa que o Com_add, mas nao da commit.
          Usar se for adicionar muitos itens.
        """
        Flag = False
        mtrz = _TextoPMatriz(CamposValores)
        if(mtrz  == False ):    return Flag
        
        #1. Monta o Código SQL
        
        #Insert + nomes dos campos
        
        CodSQL = "INSERT INTO " + Tabela + "("  
        for i in range(0,len(mtrz),1):  
            CodSQL += ("", ", ")[i != 0] + mtrz[i][0]
        CodSQL += ") "
        
        #Valores a serem adicionados
        CodSQL += "VALUES(" 
        for i in range(0,len(mtrz),1):  
            CodSQL += ("", ", ")[i != 0] + mtrz[i][1].replace("¦", ";")
        CodSQL += ");"

        cursor.execute(CodSQL)
        
        return    

###ENVIANDO OS DATAFRAMES PARA O SERVIDOR###    
def FluxoToSQL(cursor, df): 

    for index, row in df.iterrows():
        
        ID_Fluxo = row['ID_Fluxo']
        ID_Projeto = row['ID_Projeto']
        BaseData = row['BaseData']
        Data = row['Data']
        Vendido = row['Vendido']
        AVender = row['AVender']
        PermutaKinea = row['PermutaKinea']
        PermutaLDI = row['PermutaLDI']
        Impostos = row['Impostos']
        CustoDeObra = row['CustodeObra']
        TaxaAdminObra = row['TaxaAdminObra']
        Carrego = row['Carrego']
        Incorporacao = row['Incorporacao']
        ComercialMarketing = row['ComercialMarketing']
        TaxaGestaoImob = row['TaxaGestaoImob']
        OutrasDespesas = row['OutrasDespesas']
        OutrasReceitas = row['OutrasReceitas']
        Terreno = row['Terreno']
        CEPAC = row['CEPAC']
        RecDespFinanceiras = row['RecDespFinanceiras']
        Aportes = row['Aportes']
        Distribuicoes = row['Distribuicoes']
        CaixaFinalEfetivo = row['CaixaFinalEfetivo']  
        SaldoAcumulado = row['SaldoNominal']
        SaldoNominal = row['SaldoAcumulado']
        SaldoReal = row['SaldoReal']
        PercentualMesVendas = row['PercentualMesVendas']
        PercentualAcumuladoVendas = row['PercentualAcumuladoVendas']                    
                       
        textoInserir = "ID_Projeto;{};ID_Fluxo;{};BaseData;'{}';Data;'{}';Vendido;{};AVender;{};PermutaKinea;{};\
                PermutaLDI;{};Impostos;{};CustodeObra;{};TaxaAdminObra;{};Carrego;{};\
                Incorporacao;{};ComercialMarketing;{};TaxaGestaoImob;{};OutrasDespesas;{};\
                OutrasReceitas;{};Terreno;{};CEPAC;{};RecDespFinanceiras;{};Aportes;{};\
                Distribuicoes;{};CaixaFinalEfetivo;{};SaldoNominal;{};SaldoAcumulado;{};SaldoReal;{};\
                PercentualMesVendas;{};PercentualAcumuladoVendas;{}".format(
                ID_Projeto, ID_Fluxo, BaseData, Data, Vendido, AVender, PermutaKinea, PermutaLDI, Impostos, CustoDeObra,
                TaxaAdminObra, Carrego, Incorporacao, ComercialMarketing,TaxaGestaoImob, OutrasDespesas, 
                OutrasReceitas, Terreno, CEPAC, RecDespFinanceiras, Aportes, Distribuicoes, CaixaFinalEfetivo, 
                SaldoAcumulado, SaldoNominal, SaldoReal, PercentualMesVendas, PercentualAcumuladoVendas)

        Com_Soft_Add(cursor,"BD_Incorporacao_Fluxo",textoInserir)
        
    return print("A base de dados BD_Incorporacao_Fluxo foi atualizada")

def InvestidoresToSQL(cursor, df): 

    for index, row in df.iterrows():

        ID_Projeto = row['ID_Projeto']
        ID_Fluxo = row['ID_Fluxo']
        Data = row['DataInvestimento']
        SCP_1 = row['SCP1']
        SCP_2 = row['SCP2']
        SCP_3 = row['SCP3']
        SCP_4 = row['SCP4']
        SCP_5 = row['SCP5']
        SCP_6 = row['SCP6']
        SCP_7 = row['SCP7']
        SCP_8 = row['SCP8']
        SCP_9 = row['SCP9']
        SCP_10 = row['SCP10']
        SCP_11 = row['SCP11']
        SCP_12 = row['SCP12']
        SCP_13 = row['SCP13']
        textoInserir = "ID_Projeto;{};ID_Fluxo;{};DataInvestimento;'{}';SCP1;{};SCP2;{};SCP3;{};SCP4;{};\
                SCP5;{};SCP6;{};SCP7;{};SCP8;{};SCP9;{};SCP10;{};SCP11;{};SCP12;{};SCP13;{}".format(
                ID_Projeto, ID_Fluxo, Data, SCP_1, SCP_2, SCP_3, SCP_4, SCP_5, SCP_6, SCP_7, SCP_8, SCP_9, SCP_10, SCP_11,
                SCP_12, SCP_13)

        Com_Soft_Add(cursor, 'BD_Incorporacao_Investidores', textoInserir)
        
    return print("A base de dados BD_Incorporacao_Investidores foi atualizada")

def IndicadoresGeraisToSQL(cursor, df): 

    for index, row in df.iterrows():

        ID_Projeto = row['ID_Projeto']
        ID_Fluxo = row['ID_Fluxo']
        DataBaseTerreno = row['DataBaseTerreno']
        DataBaseObra = row['DataBaseObra']
        ValorMetroTerreno = row['ValorMetroTerreno']
        ValorMetroObra = row['ValorMetroObra']
        AreadoTerreno = row['AreadoTerreno']
        ValordoTerreno = row['ValordoTerreno']
        textoInserir = "ID_Projeto;{};ID_Fluxo;{};DataBaseTerreno;'{}';DataBaseObra;'{}';ValorMetroTerreno;{};\
        ValorMetroObra;{};AreadoTerreno;{};ValordoTerreno;{}".format(
                ID_Projeto,
                ID_Fluxo,
                DataBaseTerreno,
                DataBaseObra,
                ValorMetroTerreno,
                ValorMetroObra,
                AreadoTerreno,
                ValordoTerreno)

        Com_Soft_Add(cursor,"BD_Incorporacao_IndicadoresGerais",textoInserir)
        
    return print("A base de dados BD_Incorporacao_IndicadoresGerais foi atualizada")

def UnidadesToSQL(cursor, df): 
    
    for index, row in df.iterrows():

        ID_Unidade = row['ID_Unidade']
        ID_Andar = row['ID_Andar']
        ID_Tipo = row['ID_Tipo']
        ID_Classificacao = row['ID_Classificacao']
        NumeroUnidade = row['NúmeroUnidade']
        Area = row['Área']
        Status = row['Status']
        DataVenda = row['DataVenda']
        MesVenda = row['MesVenda']
        ValorVendidoOriginal = row['ValorVendidoOriginal']
        PreçoPorMetroQVendido = row['PreçoPorMetroQVendido']
        textoInserir = "ID_Unidade;{};ID_Andar;{};ID_Tipo;{};ID_Classificacao;{};NúmeroUnidade;'{}';\
                Área;{};Status;'{}';DataVenda;'{}';MesVenda;'{}';ValorVendidoOriginal;{};PreçoPorMetroQVendido;{}".format(
                ID_Unidade, ID_Andar, ID_Tipo, ID_Classificacao, NumeroUnidade, Area, Status, DataVenda, MesVenda,
                ValorVendidoOriginal, PreçoPorMetroQVendido)

        Com_Soft_Add(cursor,"BD_Incorporacao_Unidades",textoInserir)
        
    return print("A base de dados BD_Incorporacao_Unidades foi atualizada")

def BlocoToSQL(cursor, df): 

    for index, row in df.iterrows():

        ID_Projeto = row['ID_Projeto']
        ID_Bloco = row['ID_Bloco']
        NomeBloco = row['NomeBloco']
        textoInserir = "ID_Projeto;{};ID_Bloco;{};NomeBloco;'{}'".format(ID_Projeto, ID_Bloco, NomeBloco)

        Com_Soft_Add(cursor,"BD_Incorporacao_Bloco",textoInserir)
        
    return print("A base de dados BD_Incorporacao_Bloco foi atualizada")

def AndarToSQL(cursor, df): 

    for index, row in df.iterrows():

        ID_Andar = row['ID_Andar']
        ID_Bloco = row['ID_Bloco']
        Andar = row['Andar']
        textoInserir = "ID_Andar;{};ID_Bloco;{};Andar;{}".format(ID_Andar, ID_Bloco, Andar)

        Com_Soft_Add(cursor,"BD_Incorporacao_Andar",textoInserir)
        
    return print("A base de dados BD_Incorporacao_Andar foi atualizada")

def IndicadoresMensaisToSQL(cursor, df): 

    for index, row in df.iterrows():

        ID_Fluxo = row['ID_Fluxo']
        TIRReal = row['TIRReal']
        TIRNominal = row['TIRNominal']
        Inadimplencia = row['Inadimplência']
        ValorMetroHistorico = row['ValorMetroHistórico']
        ValorMetroEstoque = row['ValorMetroEstoque']
        Lucro = row['Lucro']
        LTV = row['LTV']
        PercentualAvancoObra = row['PercentualAvançoObra']
        BD_Incorporacao_Fluxo_Data = row['BD_Incorporacao_Fluxo_Data']                               
        Caixa = row['Caixa']
        Receita = row['Receita']
        ReceitaPermuta = row['ReceitaPermuta']
        Impostos = row['Impostos']
        CustodeObra = row['CustodeObra']
        TaxaAdminObra = row['TaxaAdminObra']
        Carrego = row['Carrego']
        Incorporacao = row['Incorporacao']
        Marketing = row['Marketing']
        TaxaGestaoImob = row['TaxaGestaoImob']
        OutrasDespesas = row['OutrasDespesas']
        TotalDespesas = row['TotalDespesas']
        TerrenoCEPAC = row['TerrenoCEPAC']
        RecDespFinanceiras = row['RecDespFinanceiras']
        NAVTotal = row['NAVTotal']
        NAVSCP = row['NAVSCP']
        TIRRealProjeto = row['TIRRealProjeto']
        TIRNominalProjeto = row['TIRNominalProjeto']
        TIRRealHolding = row['TIRRealHolding']
        TIRRealHoldingPosPerf = row['TIRRealHoldingPosPerf']
        TIRNominalHolding = row['TIRNominalHolding']
        Multiplo = row['Multiplo']
        Performance = row['Performance']
        textoInserir = "ID_Fluxo;{};TIRReal;{};TIRNominal;{};Inadimplência;{};ValorMetroHistórico;{};\
                        ValorMetroEstoque;{};Lucro;{};LTV;{};PercentualAvançoObra;{};BD_Incorporacao_Fluxo_Data;'{}';\
                        Caixa;{};Receita;{};ReceitaPermuta;{};Impostos;{};CustodeObra;{};TaxaAdminObra;{};Carrego;{};\
                        Incorporacao;{};Marketing;{};TaxaGestaoImob;{};OutrasDespesas;{};TotalDespesas;{};TerrenoCEPAC;{};\
                        RecDespFinanceiras;{};NAVTotal;{};NAVSCP;{};TIRRealProjeto;{};TIRNominalProjeto;{};TIRRealHolding;{};\
                        TIRRealHoldingPosPerf;{};TIRNominalHolding;{};Multiplo;{};Performance;{}".format(ID_Fluxo, 
                        TIRReal, TIRNominal, Inadimplencia, ValorMetroHistorico, ValorMetroEstoque, Lucro, LTV, 
                        PercentualAvancoObra, BD_Incorporacao_Fluxo_Data, Caixa, Receita, ReceitaPermuta, 
                        Impostos, CustodeObra, TaxaAdminObra, Carrego, Incorporacao, Marketing, TaxaGestaoImob, 
                        OutrasDespesas, TotalDespesas, TerrenoCEPAC, RecDespFinanceiras, NAVTotal, NAVSCP, 
                        TIRRealProjeto, TIRNominalProjeto, TIRRealHolding, TIRRealHoldingPosPerf, TIRNominalHolding, 
                        Multiplo, Performance)

        Com_Soft_Add(cursor,"BD_Incorporacao_IndicadoresMensais",textoInserir)
        
    return print("A base de dados BD_Incorporacao_IndicadoresMensais foi atualizada")

def BandeiraCoToSQL(cursor, df):

    for index, row in df.iterrows():

        ID_Projeto = row['ID_Projeto']        
        ID_Fluxo = row['ID_Fluxo']
        BaseData = row['BaseData']
        Data = row['Data']
        ChamadaEfetiva = row['ChamadaEfetiva']
        DistribuicaoEfetiva = row['DistribuicaoEfetiva']
        ChamadaEquity = row['ChamadaEquity']
        ChamadaPermuta = row['ChamadaPermuta']
        SPEBandeiraIncorporacaoEquity = row['SPEBandeiraIncorporacaoEquity']
        SPEBandeiraIncorporacaoPermuta = row['SPEBandeiraIncorporacaoPermuta']
        SPEBandeiraIncorporacao = row['SPEBandeiraIncorporacao']
        Impostos = row['Impostos']
        Auditoria = row['Auditoria']
        MonitoramentoObra = row['MonitoramentoObra']
        Juridico = row['Juridico']
        TaxaAdministracao = row['TaxaAdministracao']
        TaxaPerfomancePermutas = row['TaxaPerfomancePermutas']
        Contabilidade = row['Contabilidade']
        OutrasDespesas = row['OutrasDespesas']
        OutrasReceitas = row['OutrasReceitas']
        SaldoNominal = row['SaldoNominal']
        SaldoAcumulado = row['SaldoAcumulado']
        SaldoReal = row['SaldoReal']  
        SaldoRealAcumulado = row['SaldoRealAcumulado']
        BasePerformance = row['BasePerformance']
        Performance = row['Performance']                          
        SaldoNominalPosPerformance = row['SaldoNominalPosPerformance']
        SaldoAcumuladoNominalPosPerformance = row['SaldoAcumuladoNominalPosPerformance']
        SaldoRealPosPerformance = row['SaldoRealPosPerformance']                       
        textoInserir = "ID_Projeto;{};ID_Fluxo;{};BaseData;'{}';Data;'{}';ChamadaEfetiva;{};DistribuicaoEfetiva;{};ChamadaEquity;{};\
                        ChamadaPermuta;{};SPEBandeiraIncorporacaoEquity;{};SPEBandeiraIncorporacaoPermuta;{};\
                        SPEBandeiraIncorporacao;{};Impostos;{};Auditoria;{};MonitoramentoObra;{};Juridico;{};\
                        TaxaAdministracao;{};TaxaPerfomancePermutas;{};Contabilidade;{};OutrasDespesas;{};\
                        OutrasReceitas;{};SaldoNominal;{};SaldoAcumulado;{};SaldoReal;{};SaldoRealAcumulado;{};\
                        BasePerformance;{};Performance;{};SaldoNominalPosPerformance;{};\
                        SaldoAcumuladoNominalPosPerformance;{};SaldoRealPosPerformance;{}".format(
                ID_Projeto, ID_Fluxo, BaseData, Data, ChamadaEfetiva, DistribuicaoEfetiva, ChamadaEquity, ChamadaPermuta, 
                SPEBandeiraIncorporacaoEquity, SPEBandeiraIncorporacaoPermuta, SPEBandeiraIncorporacao, Impostos, Auditoria, 
                MonitoramentoObra, Juridico,TaxaAdministracao, TaxaPerfomancePermutas, Contabilidade, OutrasDespesas, OutrasReceitas, 
                SaldoNominal, SaldoAcumulado, SaldoReal, SaldoRealAcumulado, BasePerformance, Performance, 
                SaldoNominalPosPerformance, SaldoAcumuladoNominalPosPerformance, SaldoRealPosPerformance)

        Com_Soft_Add(cursor,"BD_Incorporacao_Investidores_BandeiraCo",textoInserir)

    return print("A base de dados BD_Incorporacao_Investidores_BandeiraCo foi atualizada")

###ENVIAR TODOS OS DATAFRAMES PARA O SERVIDOR###
def Update_Server():
    try:
        try:
            FluxoToSQL(cursor, DataFrame_Fluxo())
        except:
            print("Ocorreu um problema ao enviar o dataframe Fluxo") 
        try:
            InvestidoresToSQL(cursor, DataFrame_Investidores())
        except:
            print("Ocorreu um problema ao enviar o dataframe Investidores") 
        try:
            IndicadoresGeraisToSQL(cursor, DataFrame_IndicadoresGerais())
        except:
            print("Ocorreu um problema ao enviar o dataframe IndicadoresGerais") 
        try:
            UnidadesToSQL(cursor, DataFrame_Unidades())
        except:
            print("Ocorreu um problema ao enviar o dataframe Unidades")             
        try:
            BlocoToSQL(cursor, DataFrame_Bloco())
        except:
            print("Ocorreu um problema ao enviar o dataframe Bloco")             
        try:
            AndarToSQL(cursor, DataFrame_Andar())
        except:
            print("Ocorreu um problema ao enviar o dataframe Andar") 
        try:
            IndicadoresMensaisToSQL(cursor, DataFrame_IndicadoresMensais())
        except:
            print("Ocorreu um problema ao enviar o dataframe IndicadoresMensais")
        try:
            BandeiraCoToSQL(cursor, DataFrame_BandeiraCo())
        except:
            print("Ocorreu um problema ao enviar o dataframe IndicadoresMensais")
        x = "Todos os dados foram atualizados no servidor"          
    except:
        x = "Ocorreu um problema atualizar os dados no servidor"
    
    return print('{}'.format(x))


# =============================================================================
# Update_DataBase()
# =============================================================================

Update_Server()
cnxn.commit()


wb.close()
wb_Fechamento.close()

t1_stop = process_time() 
print("O programa demorou:", t1_stop-t1_start)

