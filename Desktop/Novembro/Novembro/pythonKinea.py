from datetime import datetime
from openpyxl import load_workbook
from time import process_time 
import ConvertExtension
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

t1_start = process_time() 

wb = load_workbook(filename = "planilhaIncorporadoraConverted.xlsx", read_only=True, data_only=True)
wb1 = load_workbook(filename = "planilhaSocio.xlsx", read_only=True, data_only=True)
wb_destiny = load_workbook(filename = "planilhaPrincipal.xlsx")

sheet_Unidades = wb['UNIDADES']
sheet_FluxoDeCaixa = wb['Contas a Receber']
sheet_PosFin = wb['POSIÇÃO FINANCEIRA']
sheet_Liquidacoes = wb['LIQUIDAÇÕES']
sheet_LDI = wb1['Fluxo (Socio)']

sheet_M_Premissas = wb_destiny['Premissas']
sheet_M_unidades = wb_destiny['M-Unidades']
sheet_M_FluxoDeCaixa = wb_destiny['M-FLUXO DE CAIXA']
sheet_M_PosFin = wb_destiny['M-POS. FIN.']
sheet_M_Liquidacoes = wb_destiny['M-LIQUIDAÇÕES']
sheet_M_LDI = wb_destiny['LDI - Fluxo de Caixa']

lista_unidades = []
lista_fluxoDeCaixa = []
lista_posFin = []
lista_liquidacoes = []
lista_LDI_UP = []
lista_LDI_DOWN = []
lista_emBranco = []
lista_LDI_Terreno = []

def lerArquivo(sheet, lista, row_min, row_max, col_min, col_max):
        
    if (row_max == 0):
        m_row = sheet.max_row 
    else:
        m_row = row_max
        
    for row in sheet.iter_rows(min_row=row_min, min_col=col_min, max_row=m_row, max_col=col_max):
        for cell in row:
            try:
                pass
            except:
                pass
            lista.append(cell.value)

    return lista

def enviarArquivo(sheet_destiny, lista, row_min, col_min, col_max):
    
    cont=0
    for i in range (row_min-1, (len(lista)//col_max)+row_min-1): 
        for j in range (col_max):
            try:
                sheet_destiny.cell(row = i+1, column = col_min+j).value = lista[cont]
            except:
                pass
            cont+=1
           
    return

def lerArquivoTodos():
    
    lerArquivo(sheet_Unidades, lista_unidades, 1, 0, 1, 3)
    lerArquivo(sheet_FluxoDeCaixa, lista_fluxoDeCaixa, 1, 0, 1, 14)
    lerArquivo(sheet_PosFin, lista_posFin, 1, 0, 1, 48)
    lerArquivo(sheet_Liquidacoes, lista_liquidacoes, 1, 0, 1, 6)
    lerArquivo(sheet_LDI, lista_LDI_UP, 1, 11, 1, 50)  
    lerArquivo(sheet_LDI, lista_LDI_DOWN, 154, 0, 1, 50)   
    lerArquivo(sheet_LDI, lista_LDI_Terreno, 154, 0, 11, 11)
    
    return (print("Todos os arquivos foram lidos"))

def enviarArquivoTodos():
    
    enviarArquivo(sheet_M_unidades, lista_unidades, 1, 1, 3)
    enviarArquivo(sheet_M_FluxoDeCaixa, lista_fluxoDeCaixa, 1, 1, 14)
    enviarArquivo(sheet_M_PosFin, lista_posFin, 1, 8, 48)
    enviarArquivo(sheet_M_Liquidacoes, lista_liquidacoes, 1, 8, 6)
    enviarArquivo(sheet_M_LDI, lista_LDI_UP, 8, 1, 50)   
    enviarArquivo(sheet_M_LDI, lista_LDI_DOWN, 161, 1, 50)
    enviarArquivoTerreno()
    
    return (print("Todos os arquivos foram enviados")) 

def apagarLinhasColunasLDI():
    
    m_row = sheet_M_LDI.max_row 
   
    for i in range (14, m_row): 
        sheet_M_LDI.cell(row = i+1, column = 4).value = ""
            
    for i in range (4, 53): 
        sheet_M_LDI.cell(row = 9, column = i+1).value = ""

    for i in range (6, 53): 
        sheet_M_LDI.cell(row = 8, column = i+1).value = ""
        
    for i in range (4, 6): 
        sheet_M_LDI.cell(row = 17, column = i+1).value = ""
    
    for i in range (38, 43): 
        sheet_M_LDI.cell(row = 17, column = i+1).value = ""
    
    for i in range (43,47):
        for j in range(161,194):
            sheet_M_LDI.cell(row = j, column = i).value = ""
        
    return

def enviarArquivoTerreno():
    cont1 = 0
    cont2 = 161
    row_min = 161
    col = 52
    
    for i in range (len(lista_LDI_Terreno)):
        lista_LDI_Terreno[i] = "={} - BA{}".format(lista_LDI_Terreno[i], cont2)
        cont2 += 1

    for i in range (row_min, len(lista_LDI_Terreno)+row_min-1):
        sheet_M_LDI.cell(row = i, column = col).value = lista_LDI_Terreno[cont1]
        cont1 += 1    
        
    return

def tratamentoFluxoDeCaixa():

    row_min = 7    
    cont = 0
    m_row_Incorporadora = sheet_FluxoDeCaixa.max_row
    m_row_Principal = sheet_M_FluxoDeCaixa.max_row
    lista_unidade = []
    lista_unidade2 = []
    lista_unidade3 = []
    
    for i in range (row_min, m_row_Incorporadora):
        lista_unidade.append(sheet_M_FluxoDeCaixa.cell(row = i, column = 2).value)
        lista_unidade2.append(lista_unidade[cont].split("A"))
        lista_unidade3.append(lista_unidade2[cont][1])
        sheet_M_FluxoDeCaixa.cell(row = i, column = 16).value = int(lista_unidade3[cont])
        cont+=1
        
    lista_data = []    
    lista_data2 = []
    lista_data3 = []   
    
    cont = 0
    for i in range (row_min, m_row_Incorporadora):
        lista_data.append(sheet_M_FluxoDeCaixa.cell(row = i, column = 1).value)
        lista_data2.append("01/{}/{}".format(lista_data[cont].strftime("%m"),lista_data[cont].strftime("%y")))
        lista_data3.append(datetime.strptime(lista_data2[cont], "%d/%m/%y"))
        sheet_M_FluxoDeCaixa.cell(row = i, column = 18).value = lista_data3[cont]
        cont += 1        

    for i in range (row_min, m_row_Incorporadora):
        sheet_M_FluxoDeCaixa.cell(row = i, column = 19).value = '=M{}'.format(i)
        
    dataBase = sheet_M_Premissas.cell(row = 2, column = 3).value

    for i in range (row_min, m_row_Incorporadora):
        
        data = sheet_M_FluxoDeCaixa.cell(row = i, column = 1).value
        
        if data.year <= dataBase.year:
            if data.month <= dataBase.month:
                sheet_M_FluxoDeCaixa.cell(row = i, column = 20).value = 1
            else:
                sheet_M_FluxoDeCaixa.cell(row = i, column = 20).value = 0
        else:
            sheet_M_FluxoDeCaixa.cell(row = i, column = 20).value = 0
    
    if m_row_Incorporadora < m_row_Principal:
        for i in range (m_row_Incorporadora+1, m_row_Principal+1):
            for j in range (1,20):
                sheet_M_FluxoDeCaixa.cell(row = i, column = j).value = ""
        for i in range (m_row_Incorporadora, m_row_Principal+1):
            for k in range (16,21):   
                sheet_M_FluxoDeCaixa.cell(row = i, column = k).fill = PatternFill(bgColor="000000")
                sheet_M_FluxoDeCaixa.cell(row = i, column = k).value = ""
                sheet_M_FluxoDeCaixa.cell(row = i, column = k).border = Border(left=Side(border_style=None),right=Side(border_style=None))
    else:
        pass
        
    return

lerArquivoTodos()
enviarArquivoTodos()
apagarLinhasColunasLDI()
tratamentoFluxoDeCaixa()

wb_destiny.save(filename = 'planilhaPrincipal.xlsx')
wb.close()
wb1.close()


t1_stop = process_time() 
print("O programa demorou:", t1_stop-t1_start)



